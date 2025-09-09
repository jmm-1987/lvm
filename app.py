import os
import sqlite3
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, Response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import csv
import io
import zipfile
import tempfile
import shutil
import paramiko
import re
import PyPDF2
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = 'pon_aqui_una_clave_secreta_larga_y_unica'

# Configuración de la base de datos SQLite
db_path = os.path.join(os.path.dirname(__file__), 'app.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Añadir datetime al contexto global de Jinja2
@app.context_processor
def inject_datetime():
    return dict(datetime=datetime)

# Modelo de Cuenta
class Cuenta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cuenta = db.Column(db.String(50), nullable=False)  # Ahora almacena el número completo
    nombre = db.Column(db.String(100), nullable=False)
    tipo = db.Column(db.String(20), nullable=False, default='normal')  # 'normal' o 'contrapartida'
    anotaciones = db.Column(db.String(255), nullable=True)
    cuenta_asociada_id = db.Column(db.Integer, db.ForeignKey('cuenta.id'), nullable=True)
    cuenta_asociada = db.relationship('Cuenta', remote_side=[id])

# Modelo de Movimiento
class Movimiento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(20), nullable=False)
    fecha_trabajo = db.Column(db.String(20), nullable=False)
    fecha_factura = db.Column(db.String(20), nullable=False)
    num_factura = db.Column(db.String(50), nullable=False)
    base_imponible = db.Column(db.Float, nullable=False)
    total = db.Column(db.Float, nullable=False)

# Modelo de MovimientoConcepto (relación muchos a muchos entre Movimiento y Cuenta)
class MovimientoConcepto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    movimiento_id = db.Column(db.Integer, db.ForeignKey('movimiento.id'), nullable=False)
    cuenta_id = db.Column(db.Integer, db.ForeignKey('cuenta.id'), nullable=False)
    importe = db.Column(db.Float, nullable=False)
    concepto = db.Column(db.String(100), nullable=True)
    contrapartida_id = db.Column(db.Integer, db.ForeignKey('cuenta.id'), nullable=True) # Nuevo campo para la contrapartida
    
    cuenta = db.relationship('Cuenta', foreign_keys=[cuenta_id])
    contrapartida = db.relationship('Cuenta', foreign_keys=[contrapartida_id]) # Relación con la cuenta de contrapartida

Movimiento.conceptos = db.relationship('MovimientoConcepto', backref='movimiento', cascade='all, delete-orphan')

# Modelo de Vehículo
class Vehiculo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    matricula = db.Column(db.String(20), nullable=False, unique=True)
    marca = db.Column(db.String(50), nullable=False)
    modelo = db.Column(db.String(50), nullable=False)
    año_compra = db.Column(db.Integer, nullable=False)
    fecha_alta = db.Column(db.String(20), nullable=False, default=lambda: datetime.now().strftime('%Y-%m-%d'))
    activo = db.Column(db.Boolean, nullable=False, default=True)
    observaciones = db.Column(db.String(255), nullable=True)

# Modelo de Consumo de Gasoil
class ConsumoGasoil(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    vehiculo_id = db.Column(db.Integer, db.ForeignKey('vehiculo.id'), nullable=False)
    fecha = db.Column(db.String(20), nullable=False)  # Formato YYYY-MM-DD
    litros = db.Column(db.Float, nullable=False)
    precio_litro = db.Column(db.Float, nullable=False)
    total = db.Column(db.Float, nullable=False)  # litros * precio_litro
    kms = db.Column(db.Float, nullable=False)
    facturacion = db.Column(db.Float, nullable=True)  # Ingresos generados por el vehículo
    observaciones = db.Column(db.String(255), nullable=True)
    
    vehiculo = db.relationship('Vehiculo', backref='consumos')

# Eliminar el modelo Usuario y la tabla de usuarios
# Definir un usuario en memoria para Flask-Login
class UsuarioFalso(UserMixin):
    def __init__(self, id):
        self.id = id
        self.username = 'lvm'

# Configuración de Flask-Login
login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    if user_id == 'lvm':
        return UsuarioFalso('lvm')
    return None

# Ruta de login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        #lvm_password = os.environ.get('LVM_PASSWORD')
        lvm_password = "1"
        if username == 'lvm' and lvm_password and password == lvm_password:
            user = UsuarioFalso('lvm')
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    return render_template('login.html')

# Ruta de logout
@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    subir_db_a_ftp()  # Subir la base de datos antes de cerrar sesión
    logout_user()
    return redirect(url_for('login'))

# Proteger todas las rutas excepto login y static
@app.before_request
def require_login():
    if request.endpoint not in ('login', 'static') and not current_user.is_authenticated:
        return redirect(url_for('login'))

@app.route('/')
def index():
    return render_template('index.html')

# Vistas para cuentas
@app.route('/cuentas')
def listar_cuentas():
    cuentas_normales = Cuenta.query.filter_by(tipo='normal').all()
    cuentas_contrapartida = Cuenta.query.filter_by(tipo='contrapartida').all()
    return render_template('cuentas.html', cuentas_normales=cuentas_normales, cuentas_contrapartida=cuentas_contrapartida)

@app.route('/cuentas/nueva', methods=['GET', 'POST'])
def nueva_cuenta():
    cuentas_normales = Cuenta.query.filter_by(tipo='normal').all()
    if request.method == 'POST':
        cuenta = request.form['cuenta']
        nombre = request.form['nombre']
        tipo = request.form['tipo']
        anotaciones = request.form.get('anotaciones', '')
        cuenta_asociada_id = request.form.get('cuenta_asociada_id') if tipo == 'contrapartida' else None
        nueva = Cuenta(cuenta=cuenta, nombre=nombre, tipo=tipo, anotaciones=anotaciones, cuenta_asociada_id=cuenta_asociada_id)
        db.session.add(nueva)
        db.session.commit()
        return redirect(url_for('listar_cuentas'))
    return render_template('cuenta_form.html', cuentas_normales=cuentas_normales)

@app.route('/cuentas/editar/<int:id>', methods=['GET', 'POST'])
def editar_cuenta(id):
    cuenta = Cuenta.query.get_or_404(id)
    cuentas_normales = Cuenta.query.filter_by(tipo='normal').all()
    if request.method == 'POST':
        cuenta.cuenta = request.form['cuenta']
        cuenta.nombre = request.form['nombre']
        cuenta.tipo = request.form['tipo']
        cuenta.anotaciones = request.form.get('anotaciones', '')
        cuenta.cuenta_asociada_id = request.form.get('cuenta_asociada_id') if cuenta.tipo == 'contrapartida' else None
        db.session.commit()
        return redirect(url_for('listar_cuentas'))
    return render_template('cuenta_form.html', cuenta=cuenta, cuentas_normales=cuentas_normales)

@app.route('/cuentas/borrar/<int:id>', methods=['POST'])
def borrar_cuenta(id):
    cuenta = Cuenta.query.get_or_404(id)
    # Comprobar si la cuenta está asociada a algún concepto de movimiento
    conceptos = MovimientoConcepto.query.filter((MovimientoConcepto.cuenta_id == id) | (MovimientoConcepto.contrapartida_id == id)).first()
    if conceptos:
        flash('No se puede borrar la cuenta porque está asociada a movimientos.', 'error')
        return redirect(url_for('listar_cuentas'))
    db.session.delete(cuenta)
    db.session.commit()
    flash('Cuenta borrada correctamente.', 'success')
    return redirect(url_for('listar_cuentas'))

# Vistas para movimientos
@app.route('/movimientos')
def listar_movimientos():
    # Ordenar por fecha de factura convertida a fecha real (de más reciente a más antigua)
    movimientos = db.session.query(Movimiento).order_by(
        db.func.strftime('%Y-%m-%d', 
            db.func.substr(Movimiento.fecha_factura, 7, 4) + '-' + 
            db.func.substr(Movimiento.fecha_factura, 4, 2) + '-' + 
            db.func.substr(Movimiento.fecha_factura, 1, 2)
        ).desc()
    ).all()
    conceptos_por_mov = {}
    contrapartida_por_mov = {}
    for mov in movimientos:
        conceptos = MovimientoConcepto.query.filter_by(movimiento_id=mov.id).all()
        conceptos_por_mov[mov.id] = conceptos
        # Tomar la contrapartida de la primera línea (todas deben ser iguales)
        if conceptos:
            contrapartida_id = getattr(conceptos[0], 'contrapartida_id', None)
            contrapartida = Cuenta.query.get(contrapartida_id) if contrapartida_id else None
        else:
            contrapartida = None
        contrapartida_por_mov[mov.id] = contrapartida
    # Calcular fechas por defecto
    hoy = datetime.today()
    mes = hoy.month
    if mes <= 3:
        inicio_trimestre = datetime(hoy.year, 1, 1)
    elif mes <= 6:
        inicio_trimestre = datetime(hoy.year, 4, 1)
    elif mes <= 9:
        inicio_trimestre = datetime(hoy.year, 7, 1)
    else:
        inicio_trimestre = datetime(hoy.year, 10, 1)
    fecha_desde_default = inicio_trimestre.strftime('%Y-%m-%d')
    fecha_hasta_default = hoy.strftime('%Y-%m-%d')
    return render_template('movimientos.html', movimientos=movimientos, conceptos_por_mov=conceptos_por_mov, contrapartida_por_mov=contrapartida_por_mov, fecha_desde_default=fecha_desde_default, fecha_hasta_default=fecha_hasta_default)

@app.route('/movimientos/nuevo', methods=['GET', 'POST'])
def nuevo_movimiento():
    cuentas_normales = Cuenta.query.filter_by(tipo='normal').all()
    cuentas_contrapartida = Cuenta.query.filter_by(tipo='contrapartida').all()
    if request.method == 'POST':
        datos = request.form
        # Validar que el número de factura no esté duplicado
        num_factura = datos['num_factura']
        existe = Movimiento.query.filter_by(num_factura=num_factura).first()
        if existe:
            flash('Ya existe un movimiento con ese número de factura.', 'error')
            return render_template('movimiento_form.html', cuentas_normales=cuentas_normales, cuentas_contrapartida=cuentas_contrapartida, movimiento=None, conceptos=None)
        # Validar que todas las contrapartidas sean iguales
        contrapartidas = set()
        idx = 0
        while f'contrapartida_{idx}' in datos:
            contrapartidas.add(datos[f'contrapartida_{idx}'])
            idx += 1
        if len(contrapartidas) > 1:
            flash('Todas las líneas deben tener la misma cuenta de contrapartida.', 'error')
            return render_template('movimiento_form.html', cuentas_normales=cuentas_normales, cuentas_contrapartida=cuentas_contrapartida, movimiento=None, conceptos=None)
        # Guardar el movimiento
        nuevo = Movimiento(
            tipo=datos['tipo'],
            fecha_trabajo=datos['fecha_trabajo'],
            fecha_factura=datos['fecha_factura'],
            num_factura=datos['num_factura'],
            base_imponible=float(datos['base_imponible']) if 'base_imponible' in datos else 0,
            total=float(datos['total'])
        )
        db.session.add(nuevo)
        db.session.flush()
        idx = 0
        while f'cuenta_{idx}' in datos:
            cuenta_id = int(datos[f'cuenta_{idx}'])
            contrapartida_id = int(datos[f'contrapartida_{idx}'])
            importe = float(datos[f'importe_{idx}'])
            concepto = datos.get(f'concepto_{idx}', '')
            concepto_obj = MovimientoConcepto(movimiento_id=nuevo.id, cuenta_id=cuenta_id, importe=importe, concepto=concepto)
            # Guardar la contrapartida como campo adicional si lo necesitas
            concepto_obj.contrapartida_id = contrapartida_id
            db.session.add(concepto_obj)
            idx += 1
        db.session.commit()
        return redirect(url_for('listar_movimientos'))
    return render_template('movimiento_form.html', cuentas_normales=cuentas_normales, cuentas_contrapartida=cuentas_contrapartida)

@app.route('/movimientos/editar/<int:id>', methods=['GET', 'POST'])
def editar_movimiento(id):
    movimiento = Movimiento.query.get_or_404(id)
    cuentas_normales = Cuenta.query.filter_by(tipo='normal').all()
    cuentas_contrapartida = Cuenta.query.filter_by(tipo='contrapartida').all()
    if request.method == 'POST':
        datos = request.form
        # Validar que el número de factura no esté duplicado (excepto el propio movimiento)
        num_factura = datos['num_factura']
        existe = Movimiento.query.filter(Movimiento.num_factura == num_factura, Movimiento.id != movimiento.id).first()
        if existe:
            flash('Ya existe un movimiento con ese número de factura.', 'error')
            conceptos = MovimientoConcepto.query.filter_by(movimiento_id=movimiento.id).all()
            return render_template('movimiento_form.html', movimiento=movimiento, cuentas_normales=cuentas_normales, cuentas_contrapartida=cuentas_contrapartida, conceptos=conceptos)
        # Validar que todas las contrapartidas sean iguales
        contrapartidas = set()
        idx = 0
        while f'contrapartida_{idx}' in datos:
            contrapartidas.add(datos[f'contrapartida_{idx}'])
            idx += 1
        if len(contrapartidas) > 1:
            flash('Todas las líneas deben tener la misma cuenta de contrapartida.', 'error')
            conceptos = MovimientoConcepto.query.filter_by(movimiento_id=movimiento.id).all()
            return render_template('movimiento_form.html', movimiento=movimiento, cuentas_normales=cuentas_normales, cuentas_contrapartida=cuentas_contrapartida, conceptos=conceptos)
        movimiento.tipo = datos['tipo']
        movimiento.fecha_trabajo = datos['fecha_trabajo']
        movimiento.fecha_factura = datos['fecha_factura']
        movimiento.num_factura = datos['num_factura']
        movimiento.base_imponible = float(datos['base_imponible']) if 'base_imponible' in datos else 0
        movimiento.total = float(datos['total'])
        MovimientoConcepto.query.filter_by(movimiento_id=movimiento.id).delete()
        idx = 0
        while f'cuenta_{idx}' in datos:
            cuenta_id = int(datos[f'cuenta_{idx}'])
            contrapartida_id = int(datos[f'contrapartida_{idx}'])
            importe = float(datos[f'importe_{idx}'])
            concepto = datos.get(f'concepto_{idx}', '')
            concepto_obj = MovimientoConcepto(movimiento_id=movimiento.id, cuenta_id=cuenta_id, importe=importe, concepto=concepto)
            concepto_obj.contrapartida_id = contrapartida_id
            db.session.add(concepto_obj)
            idx += 1
        db.session.commit()
        return redirect(url_for('listar_movimientos'))
    conceptos = MovimientoConcepto.query.filter_by(movimiento_id=movimiento.id).all()
    return render_template('movimiento_form.html', movimiento=movimiento, cuentas_normales=cuentas_normales, cuentas_contrapartida=cuentas_contrapartida, conceptos=conceptos)

@app.route('/movimientos/borrar/<int:id>', methods=['POST'])
def borrar_movimiento(id):
    movimiento = Movimiento.query.get_or_404(id)
    db.session.delete(movimiento)
    db.session.commit()
    flash('Movimiento borrado correctamente.', 'success')
    return redirect(url_for('listar_movimientos'))

@app.route('/movimientos/duplicar/<int:id>', methods=['GET'])
@login_required
def duplicar_movimiento(id):
    movimiento = Movimiento.query.get_or_404(id)
    conceptos = MovimientoConcepto.query.filter_by(movimiento_id=movimiento.id).all()
    cuentas_normales = Cuenta.query.filter_by(tipo='normal').all()
    cuentas_contrapartida = Cuenta.query.filter_by(tipo='contrapartida').all()
    # Preparamos un objeto similar al de edición pero con fechas y factura vacías
    movimiento_clon = Movimiento(
        tipo=movimiento.tipo,
        fecha_trabajo='',
        fecha_factura='',
        num_factura='',
        base_imponible=movimiento.base_imponible,
        total=movimiento.total
    )
    # Pasamos los conceptos para que se muestren en el formulario
    return render_template('movimiento_form.html', movimiento=movimiento_clon, conceptos=conceptos, cuentas_normales=cuentas_normales, cuentas_contrapartida=cuentas_contrapartida)

@app.route('/movimientos/ver/<int:id>', methods=['GET'])
@login_required
def ver_movimiento(id):
    movimiento = Movimiento.query.get_or_404(id)
    conceptos = MovimientoConcepto.query.filter_by(movimiento_id=movimiento.id).all()
    cuentas_normales = Cuenta.query.filter_by(tipo='normal').all()
    cuentas_contrapartida = Cuenta.query.filter_by(tipo='contrapartida').all()
    return render_template('movimiento_form.html', movimiento=movimiento, conceptos=conceptos, cuentas_normales=cuentas_normales, cuentas_contrapartida=cuentas_contrapartida, solo_lectura=True)

@app.route('/debug_movimiento/<int:id>')
@login_required
def debug_movimiento(id):
    movimiento = Movimiento.query.get_or_404(id)
    conceptos = MovimientoConcepto.query.filter_by(movimiento_id=movimiento.id).all()
    
    debug_info = f"""
    <h2>Debug Movimiento ID: {id}</h2>
    <p><strong>Número de factura:</strong> {movimiento.num_factura}</p>
    <p><strong>Total del movimiento:</strong> {movimiento.total} €</p>
    <p><strong>Número de conceptos:</strong> {len(conceptos)}</p>
    <h3>Conceptos:</h3>
    <ul>
    """
    
    suma_conceptos = 0
    for c in conceptos:
        debug_info += f"<li>Cuenta: {c.cuenta.cuenta} - {c.cuenta.nombre} | Importe: {c.importe} € | Contrapartida: {c.contrapartida.cuenta if c.contrapartida else 'Sin contrapartida'}</li>"
        suma_conceptos += c.importe
    
    debug_info += f"""
    </ul>
    <p><strong>Suma de conceptos:</strong> {suma_conceptos} €</p>
    <p><strong>Diferencia:</strong> {movimiento.total - suma_conceptos} €</p>
    """
    
    return debug_info

@app.route('/buscar_movimiento/<factura>')
@login_required
def buscar_movimiento(factura):
    movimientos = Movimiento.query.filter_by(num_factura=factura).all()
    
    debug_info = f"""
    <h2>Búsqueda de movimientos con factura: {factura}</h2>
    <p><strong>Número de movimientos encontrados:</strong> {len(movimientos)}</p>
    """
    
    for mov in movimientos:
        conceptos = MovimientoConcepto.query.filter_by(movimiento_id=mov.id).all()
        suma_conceptos = sum(c.importe for c in conceptos)
        debug_info += f"""
        <h3>Movimiento ID: {mov.id}</h3>
        <p><strong>Total:</strong> {mov.total} €</p>
        <p><strong>Suma conceptos:</strong> {suma_conceptos} €</p>
        <p><strong>Diferencia:</strong> {mov.total - suma_conceptos} €</p>
        <p><a href="/debug_movimiento/{mov.id}">Ver detalles completos</a></p>
        <hr>
        """
    
    return debug_info

@app.route('/resultado_explotacion', methods=['GET', 'POST'])
def resultado_explotacion():
    resultado = None
    detalle = []
    diferencia = None
    # Calcular fechas por defecto
    hoy = datetime.today()
    mes = hoy.month
    if mes <= 3:
        inicio_trimestre = datetime(hoy.year, 1, 1)
    elif mes <= 6:
        inicio_trimestre = datetime(hoy.year, 4, 1)
    elif mes <= 9:
        inicio_trimestre = datetime(hoy.year, 7, 1)
    else:
        inicio_trimestre = datetime(hoy.year, 10, 1)
    fecha_inicio = inicio_trimestre.strftime('%Y-%m-%d')
    fecha_fin = hoy.strftime('%Y-%m-%d')
    resultado_explotacion = None
    if request.method == 'POST':
        fecha_inicio = request.form['fecha_inicio']
        fecha_fin = request.form['fecha_fin']
        # Buscar conceptos en ese rango de fechas
        # Convertir fechas de YYYY-MM-DD a objetos datetime para la comparación
        fecha_inicio_dt, fecha_fin_dt = convertir_fechas_para_filtro(fecha_inicio, fecha_fin)
        
        conceptos = db.session.query(MovimientoConcepto, Movimiento, Cuenta)\
            .join(Movimiento)\
            .join(Cuenta, MovimientoConcepto.cuenta_id == Cuenta.id)\
            .all()
        
        # Filtrar por fecha usando comparación de datetime
        conceptos_filtrados = []
        for c in conceptos:
            fecha_movimiento = parsear_fecha_robusto(c.Movimiento.fecha_factura)
            if fecha_movimiento and fecha_inicio_dt <= fecha_movimiento <= fecha_fin_dt:
                conceptos_filtrados.append(c)
        
        conceptos = conceptos_filtrados
        
        # Debug: imprimir información sobre los conceptos encontrados
        print(f"Conceptos encontrados en rango {fecha_inicio} a {fecha_fin}: {len(conceptos)}")
        for c in conceptos:
            print(f"Cuenta: {c.Cuenta.cuenta} - {c.Cuenta.nombre}, Importe: {c.MovimientoConcepto.importe}, Fecha: {c.Movimiento.fecha_factura}")
        
        # Buscar específicamente las nóminas
        nominas = db.session.query(MovimientoConcepto, Movimiento, Cuenta)\
            .join(Movimiento)\
            .join(Cuenta, MovimientoConcepto.cuenta_id == Cuenta.id)\
            .filter(
                Movimiento.num_factura.like('Nómina%')
            ).all()
        
        # Filtrar nóminas por fecha usando comparación de datetime
        nominas_filtradas = []
        for n in nominas:
            fecha_movimiento = parsear_fecha_robusto(n.Movimiento.fecha_factura)
            if fecha_movimiento and fecha_inicio_dt <= fecha_movimiento <= fecha_fin_dt:
                nominas_filtradas.append(n)
        
        nominas = nominas_filtradas
        
        print(f"Nóminas encontradas: {len(nominas)}")
        for n in nominas:
            print(f"Nómina: {n.Movimiento.num_factura}, Cuenta: {n.Cuenta.cuenta} - {n.Cuenta.nombre}, Importe: {n.MovimientoConcepto.importe}")
        
        # Agrupar y sumar importes por cuenta, incluyendo detalles de transacciones
        prefijos = ('623','626','621','622','625','628','629','310','640','641','642','649','662','7')
        cuentas_excluidas = ('642000000002',)  # Cuentas específicas a excluir del resultado de explotación
        cuentas_resultado_neto = ('74000000002', '76900000001')  # Cuentas que se suman al resultado neto
        detalle_dict = {}
        for c in conceptos:
            cuenta = str(c.Cuenta.cuenta)
            # Verificar si la cuenta empieza con alguno de los prefijos
            incluir_cuenta = False
            for prefijo in prefijos:
                if cuenta.startswith(prefijo):
                    incluir_cuenta = True
                    break
            
            # Excluir cuentas específicas aunque cumplan con los prefijos
            if cuenta in cuentas_excluidas or cuenta in cuentas_resultado_neto:
                incluir_cuenta = False
            
            if incluir_cuenta:
                key = cuenta
                if key not in detalle_dict:
                    detalle_dict[key] = {
                        'cuenta': cuenta,
                        'nombre': c.Cuenta.nombre,
                        'importe': 0,
                        'transacciones': []
                    }
                detalle_dict[key]['importe'] += c.MovimientoConcepto.importe
                # Añadir detalles de la transacción
                contrapartida_info = ""
                if c.MovimientoConcepto.contrapartida:
                    contrapartida_info = f"{c.MovimientoConcepto.contrapartida.cuenta} - {c.MovimientoConcepto.contrapartida.nombre}"
                else:
                    contrapartida_info = "Sin contrapartida"
                
                transaccion = {
                    'fecha': c.Movimiento.fecha_factura,
                    'contrapartida': contrapartida_info,
                    'importe': c.MovimientoConcepto.importe,
                    'concepto': c.MovimientoConcepto.concepto or "Sin concepto"
                }
                detalle_dict[key]['transacciones'].append(transaccion)
        detalle = list(detalle_dict.values())
        # Ordenar por número de cuenta
        detalle = sorted(detalle, key=lambda x: x['cuenta'])
        suma_detalle = sum(d['importe'] for d in detalle)
        suma_7 = sum(d['importe'] for d in detalle if str(d['cuenta']).startswith('7'))
        suma_705 = sum(d['importe'] for d in detalle if str(d['cuenta']).strip() == '70500000001')
        suma_resto = sum(d['importe'] for d in detalle if not str(d['cuenta']).startswith('7'))
        resultado = suma_7
        diferencia = suma_resto  # Gastos fijos (todas las cuentas que no son 7)
        resultado_explotacion = suma_705 - suma_resto
        
        # Calcular importes de las cuentas del resultado neto y agregarlas al detalle
        suma_resultado_neto = 0
        detalle_resultado_neto = []
        
        for c in conceptos:
            cuenta = str(c.Cuenta.cuenta)
            if cuenta in cuentas_resultado_neto:
                suma_resultado_neto += c.MovimientoConcepto.importe
                
                # Agregar al detalle del resultado neto
                key = cuenta
                if key not in [d['cuenta'] for d in detalle_resultado_neto]:
                    detalle_resultado_neto.append({
                        'cuenta': cuenta,
                        'nombre': c.Cuenta.nombre,
                        'importe': 0,
                        'transacciones': []
                    })
                
                # Encontrar la entrada en detalle_resultado_neto y actualizar
                for d in detalle_resultado_neto:
                    if d['cuenta'] == cuenta:
                        d['importe'] += c.MovimientoConcepto.importe
                        # Añadir detalles de la transacción
                        contrapartida_info = ""
                        if c.MovimientoConcepto.contrapartida:
                            contrapartida_info = f"{c.MovimientoConcepto.contrapartida.cuenta} - {c.MovimientoConcepto.contrapartida.nombre}"
                        else:
                            contrapartida_info = "Sin contrapartida"
                        
                        transaccion = {
                            'fecha': c.Movimiento.fecha_factura,
                            'contrapartida': contrapartida_info,
                            'importe': c.MovimientoConcepto.importe,
                            'concepto': c.MovimientoConcepto.concepto or "Sin concepto"
                        }
                        d['transacciones'].append(transaccion)
                        break
        
        # Calcular resultado neto
        resultado_neto = resultado_explotacion + suma_resultado_neto
    return render_template('resultado_explotacion.html', resultado=resultado, detalle=detalle, diferencia=diferencia, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, resultado_explotacion=resultado_explotacion, suma_resultado_neto=suma_resultado_neto, resultado_neto=resultado_neto, detalle_resultado_neto=detalle_resultado_neto)

@app.route('/iva', methods=['GET', 'POST'])
def resultado_iva():
    resultado = None
    iva_repercutido = 0
    iva_soportado = 0
    # Calcular fechas por defecto
    hoy = datetime.today()
    mes = hoy.month
    if mes <= 3:
        inicio_trimestre = datetime(hoy.year, 1, 1)
    elif mes <= 6:
        inicio_trimestre = datetime(hoy.year, 4, 1)
    elif mes <= 9:
        inicio_trimestre = datetime(hoy.year, 7, 1)
    else:
        inicio_trimestre = datetime(hoy.year, 10, 1)
    fecha_inicio = inicio_trimestre.strftime('%Y-%m-%d')
    fecha_fin = hoy.strftime('%Y-%m-%d')
    desglose_checked = False
    desglose_contrapartidas = []
    total_repercutido = 0
    total_soportado = 0
    if request.method == 'POST':
        fecha_inicio = request.form['fecha_inicio']
        fecha_fin = request.form['fecha_fin']
        desglose_checked = 'desglose' in request.form
        # Convertir fechas de YYYY-MM-DD a objetos datetime para la comparación
        fecha_inicio_dt, fecha_fin_dt = convertir_fechas_para_filtro(fecha_inicio, fecha_fin)
        
        conceptos = db.session.query(MovimientoConcepto, Movimiento, Cuenta)\
            .join(Movimiento)\
            .join(Cuenta, MovimientoConcepto.cuenta_id == Cuenta.id)\
            .filter(
                Cuenta.cuenta.in_(['47700000001', '47200000001'])
            ).all()
        
        # Filtrar por fecha usando comparación de datetime
        conceptos_filtrados = []
        for c in conceptos:
            fecha_movimiento = parsear_fecha_robusto(c.Movimiento.fecha_factura)
            if fecha_movimiento and fecha_inicio_dt <= fecha_movimiento <= fecha_fin_dt:
                conceptos_filtrados.append(c)
        
        conceptos = conceptos_filtrados
        for c in conceptos:
            if str(c.Cuenta.cuenta).strip() == '47700000001':
                iva_repercutido += c.MovimientoConcepto.importe
            elif str(c.Cuenta.cuenta).strip() == '47200000001':
                iva_soportado += c.MovimientoConcepto.importe
        resultado = iva_repercutido - iva_soportado
        # Desglose por contrapartida
        if desglose_checked:
            desglose_dict = {}
            for c in conceptos:
                contrapartida = c.MovimientoConcepto.contrapartida
                if not contrapartida:
                    continue
                key = contrapartida.cuenta
                if key not in desglose_dict:
                    desglose_dict[key] = {
                        'contrapartida': contrapartida.cuenta,
                        'nombre': contrapartida.nombre,
                        'iva_repercutido': 0,
                        'iva_soportado': 0
                    }
                if str(c.Cuenta.cuenta).strip() == '47700000001':
                    desglose_dict[key]['iva_repercutido'] += c.MovimientoConcepto.importe
                elif str(c.Cuenta.cuenta).strip() == '47200000001':
                    desglose_dict[key]['iva_soportado'] += c.MovimientoConcepto.importe
            # Filtrar solo las filas con algún valor distinto de 0
            desglose_contrapartidas = [v for v in desglose_dict.values() if v['iva_repercutido'] != 0 or v['iva_soportado'] != 0]
            # Calcular los totales
            total_repercutido = sum(v['iva_repercutido'] for v in desglose_contrapartidas)
            total_soportado = sum(v['iva_soportado'] for v in desglose_contrapartidas)
    return render_template('iva.html', resultado=resultado, iva_repercutido=iva_repercutido, iva_soportado=iva_soportado, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, desglose_checked=desglose_checked, desglose_contrapartidas=desglose_contrapartidas, total_repercutido=total_repercutido, total_soportado=total_soportado)

@app.route('/retencion_alquileres', methods=['GET', 'POST'])
def retencion_alquileres():
    # Fechas por defecto: inicio de trimestre y hoy
    hoy = datetime.today()
    mes = hoy.month
    if mes <= 3:
        inicio_trimestre = datetime(hoy.year, 1, 1)
    elif mes <= 6:
        inicio_trimestre = datetime(hoy.year, 4, 1)
    elif mes <= 9:
        inicio_trimestre = datetime(hoy.year, 7, 1)
    else:
        inicio_trimestre = datetime(hoy.year, 10, 1)
    fecha_inicio = inicio_trimestre.strftime('%Y-%m-%d')
    fecha_fin = hoy.strftime('%Y-%m-%d')
    resultado = None
    desglose_contrapartidas = []
    total_importe = 0
    if request.method == 'POST':
        fecha_inicio = request.form['fecha_inicio']
        fecha_fin = request.form['fecha_fin']
    
    # Convertir fechas de YYYY-MM-DD a objetos datetime para la comparación
    fecha_inicio_dt, fecha_fin_dt = convertir_fechas_para_filtro(fecha_inicio, fecha_fin)
    
    conceptos = db.session.query(MovimientoConcepto, Movimiento, Cuenta)\
        .join(Movimiento)\
        .join(Cuenta, MovimientoConcepto.cuenta_id == Cuenta.id)\
        .filter(
            Cuenta.cuenta == '47510000003'
        ).all()
    
    # Filtrar por fecha usando comparación de datetime
    conceptos_filtrados = []
    for c in conceptos:
        fecha_movimiento = parsear_fecha_robusto(c.Movimiento.fecha_factura)
        if fecha_movimiento and fecha_inicio_dt <= fecha_movimiento <= fecha_fin_dt:
            conceptos_filtrados.append(c)
    
    conceptos = conceptos_filtrados
    resultado = sum(c.MovimientoConcepto.importe for c in conceptos)
    # Desglose por contrapartida
    desglose_dict = {}
    for c in conceptos:
        contrapartida = c.MovimientoConcepto.contrapartida
        if not contrapartida:
            continue
        key = contrapartida.cuenta
        if key not in desglose_dict:
            desglose_dict[key] = {
                'contrapartida': contrapartida.cuenta,
                'nombre': contrapartida.nombre,
                'importe': 0
            }
        desglose_dict[key]['importe'] += c.MovimientoConcepto.importe
    desglose_contrapartidas = [v for v in desglose_dict.values() if v['importe'] != 0]
    total_importe = sum(v['importe'] for v in desglose_contrapartidas)
    return render_template('retencion_alquileres.html', resultado=resultado, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, desglose_contrapartidas=desglose_contrapartidas, total_importe=total_importe)

@app.route('/retencion_empleados', methods=['GET', 'POST'])
def retencion_empleados():
    # Fechas por defecto: inicio de trimestre y hoy
    hoy = datetime.today()
    mes = hoy.month
    if mes <= 3:
        inicio_trimestre = datetime(hoy.year, 1, 1)
    elif mes <= 6:
        inicio_trimestre = datetime(hoy.year, 4, 1)
    elif mes <= 9:
        inicio_trimestre = datetime(hoy.year, 7, 1)
    else:
        inicio_trimestre = datetime(hoy.year, 10, 1)
    fecha_inicio = inicio_trimestre.strftime('%Y-%m-%d')
    fecha_fin = hoy.strftime('%Y-%m-%d')
    resultado = None
    desglose_contrapartidas = []
    total_importe = 0
    if request.method == 'POST':
        fecha_inicio = request.form['fecha_inicio']
        fecha_fin = request.form['fecha_fin']
    
    # Convertir fechas de YYYY-MM-DD a objetos datetime para la comparación
    fecha_inicio_dt, fecha_fin_dt = convertir_fechas_para_filtro(fecha_inicio, fecha_fin)
    
    conceptos = db.session.query(MovimientoConcepto, Movimiento, Cuenta)\
        .join(Movimiento)\
        .join(Cuenta, MovimientoConcepto.cuenta_id == Cuenta.id)\
        .filter(
            Cuenta.cuenta == '47510000001'
        ).all()
    
    # Filtrar por fecha usando comparación de datetime
    conceptos_filtrados = []
    for c in conceptos:
        fecha_movimiento = parsear_fecha_robusto(c.Movimiento.fecha_factura)
        if fecha_movimiento and fecha_inicio_dt <= fecha_movimiento <= fecha_fin_dt:
            conceptos_filtrados.append(c)
    
    conceptos = conceptos_filtrados
    resultado = sum(c.MovimientoConcepto.importe for c in conceptos)
    # Desglose por contrapartida
    desglose_dict = {}
    for c in conceptos:
        contrapartida = c.MovimientoConcepto.contrapartida
        if not contrapartida:
            continue
        key = contrapartida.cuenta
        if key not in desglose_dict:
            desglose_dict[key] = {
                'contrapartida': contrapartida.cuenta,
                'nombre': contrapartida.nombre,
                'importe': 0
            }
        desglose_dict[key]['importe'] += c.MovimientoConcepto.importe
    desglose_contrapartidas = [v for v in desglose_dict.values() if v['importe'] != 0]
    total_importe = sum(v['importe'] for v in desglose_contrapartidas)
    return render_template('retencion_empleados.html', resultado=resultado, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, desglose_contrapartidas=desglose_contrapartidas, total_importe=total_importe)

@app.route('/347', methods=['GET', 'POST'])
def informe_347():
    # Fechas por defecto: inicio de trimestre y hoy
    hoy = datetime.today()
    mes = hoy.month
    if mes <= 3:
        inicio_trimestre = datetime(hoy.year, 1, 1)
    elif mes <= 6:
        inicio_trimestre = datetime(hoy.year, 4, 1)
    elif mes <= 9:
        inicio_trimestre = datetime(hoy.year, 7, 1)
    else:
        inicio_trimestre = datetime(hoy.year, 10, 1)
    fecha_inicio = inicio_trimestre.strftime('%Y-%m-%d')
    fecha_fin = hoy.strftime('%Y-%m-%d')
    resumen_contrapartidas = []
    resumen_contrapartidas_3000 = []
    resumen_contrapartidas_menos_3000 = []
    total_importe = 0
    total_importe_3000 = 0
    total_importe_menos_3000 = 0
    if request.method == 'POST':
        fecha_inicio = request.form['fecha_inicio']
        fecha_fin = request.form['fecha_fin']
    
    # Convertir fechas de YYYY-MM-DD a objetos datetime para la comparación
    fecha_inicio_dt, fecha_fin_dt = convertir_fechas_para_filtro(fecha_inicio, fecha_fin)
    
    # Agrupar por contrapartida y sumar importes
    conceptos = db.session.query(MovimientoConcepto, Movimiento, Cuenta)\
        .join(Movimiento)\
        .join(Cuenta, MovimientoConcepto.contrapartida_id == Cuenta.id)\
        .all()
    
    # Filtrar por fecha usando comparación de datetime
    conceptos_filtrados = []
    for c in conceptos:
        fecha_movimiento = parsear_fecha_robusto(c.Movimiento.fecha_factura)
        if fecha_movimiento and fecha_inicio_dt <= fecha_movimiento <= fecha_fin_dt:
            conceptos_filtrados.append(c)
    
    conceptos = conceptos_filtrados
    resumen_dict = {}
    for c in conceptos:
        contrapartida = c.Cuenta
        if not contrapartida:
            continue
        key = contrapartida.cuenta
        if key not in resumen_dict:
            resumen_dict[key] = {
                'contrapartida': contrapartida.cuenta,
                'nombre': contrapartida.nombre,
                'importe': 0
            }
        resumen_dict[key]['importe'] += c.MovimientoConcepto.importe
    resumen_contrapartidas = [v for v in resumen_dict.values() if v['importe'] != 0]
    # Ordenar alfabéticamente por nombre
    resumen_contrapartidas = sorted(resumen_contrapartidas, key=lambda x: x['nombre'])
    # Separar por umbral de 3000 euros
    for contrapartida in resumen_contrapartidas:
        if contrapartida['importe'] >= 3000:
            resumen_contrapartidas_3000.append(contrapartida)
            total_importe_3000 += contrapartida['importe']
        else:
            resumen_contrapartidas_menos_3000.append(contrapartida)
            total_importe_menos_3000 += contrapartida['importe']
    total_importe = sum(v['importe'] for v in resumen_contrapartidas)
    return render_template('347.html', fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, 
                         resumen_contrapartidas=resumen_contrapartidas, 
                         resumen_contrapartidas_3000=resumen_contrapartidas_3000,
                         resumen_contrapartidas_menos_3000=resumen_contrapartidas_menos_3000,
                         total_importe=total_importe,
                         total_importe_3000=total_importe_3000,
                         total_importe_menos_3000=total_importe_menos_3000)

@app.route('/descargar_db')
def descargar_db():
    db_path = os.path.join(os.path.dirname(__file__), 'app.db')
    fecha = datetime.now().strftime('%Y%m%d')
    nombre_archivo = f'lvm{fecha}.db'
    return send_file(db_path, as_attachment=True, download_name=nombre_archivo)

# Función para subir la base de datos por SFTP

def subir_db_a_ftp():
    sftp_host = os.environ.get('FTP_HOST')
    sftp_user = os.environ.get('FTP_USER')
    sftp_pass = os.environ.get('FTP_PASS')
    sftp_dir = os.environ.get('FTP_DIR', '/')
    db_path = os.path.join(os.path.dirname(__file__), 'app.db')
    fecha = datetime.now().strftime('%Y%m%d')
    nombre_archivo = f'lvm{fecha}.db'
    if not sftp_host or not sftp_user or not sftp_pass:
        print('Faltan variables de entorno para la conexión SFTP.')
        return
    try:
        transport = paramiko.Transport((sftp_host, 22))
        transport.connect(username=sftp_user, password=sftp_pass)
        sftp = paramiko.SFTPClient.from_transport(transport)
        # Intentar cambiar al directorio, si falla lo crea
        try:
            sftp.chdir(sftp_dir)
        except IOError:
            # Crear el directorio (soporta rutas anidadas)
            dirs = sftp_dir.strip('/').split('/')
            path = ''
            for d in dirs:
                path += '/' + d
                try:
                    sftp.chdir(path)
                except IOError:
                    sftp.mkdir(path)
                    sftp.chdir(path)
        # Listar archivos de backup existentes
        archivos = sftp.listdir()
        backups = sorted([f for f in archivos if f.startswith('lvm') and f.endswith('.db')])
        # Si hay más de 2, borrar los más antiguos (dejar solo los 2 más recientes)
        if len(backups) > 2:
            for f in backups[:-2]:
                try:
                    sftp.remove(f)
                    print(f'Backup antiguo eliminado: {f}')
                except Exception as e:
                    print(f'No se pudo eliminar {f}: {e}')
        # Subir el nuevo backup
        sftp.put(db_path, nombre_archivo)
        sftp.close()
        transport.close()
        print(f'Backup de la base de datos subido por SFTP como {nombre_archivo}.')
    except Exception as e:
        print(f'Error al subir el backup por SFTP: {e}')

def exportar_csv_a_ftp():
    sftp_host = os.environ.get('FTP_HOST')
    sftp_user = os.environ.get('FTP_USER')
    sftp_pass = os.environ.get('FTP_PASS')
    sftp_dir = os.environ.get('FTP_DIR', '/')
    fecha = datetime.now().strftime('%Y%m%d')
    nombre_archivo = f'lvm_csv_{fecha}.zip'
    
    if not sftp_host or not sftp_user or not sftp_pass:
        print('Faltan variables de entorno para la conexión SFTP.')
        return
    
    try:
        # Crear archivo ZIP con todos los CSV
        import zipfile
        zip_path = os.path.join(os.path.dirname(__file__), nombre_archivo)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Exportar cada tabla a CSV
            tablas = [Cuenta, Movimiento, MovimientoConcepto]
            nombres_tablas = ['cuentas', 'movimientos', 'movimientos_conceptos']
            
            for tabla, nombre in zip(tablas, nombres_tablas):
                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)
                
                # Obtener datos de la tabla
                registros = tabla.query.all()
                
                if registros:
                    # Escribir encabezados
                    columnas = [column.name for column in tabla.__table__.columns]
                    writer.writerow(columnas)
                    
                    # Escribir datos
                    for registro in registros:
                        fila = []
                        for columna in columnas:
                            valor = getattr(registro, columna)
                            fila.append(str(valor) if valor is not None else '')
                        writer.writerow(fila)
                
                # Añadir CSV al ZIP
                zipf.writestr(f'{nombre}.csv', csv_buffer.getvalue())
        
        # Subir ZIP al FTP
        transport = paramiko.Transport((sftp_host, 22))
        transport.connect(username=sftp_user, password=sftp_pass)
        sftp = paramiko.SFTPClient.from_transport(transport)
        
        try:
            sftp.chdir(sftp_dir)
        except IOError:
            dirs = sftp_dir.strip('/').split('/')
            path = ''
            for d in dirs:
                path += '/' + d
                try:
                    sftp.chdir(path)
                except IOError:
                    sftp.mkdir(path)
                    sftp.chdir(path)
        
        # Listar archivos de backup CSV existentes
        archivos = sftp.listdir()
        backups_csv = sorted([f for f in archivos if f.startswith('lvm_csv_') and f.endswith('.zip')])
        # Si hay más de 2, borrar los más antiguos
        if len(backups_csv) > 2:
            for f in backups_csv[:-2]:
                try:
                    sftp.remove(f)
                    print(f'Backup CSV antiguo eliminado: {f}')
                except Exception as e:
                    print(f'No se pudo eliminar {f}: {e}')
        
        # Subir el nuevo backup CSV
        sftp.put(zip_path, nombre_archivo)
        sftp.close()
        transport.close()
        
        # Eliminar archivo ZIP local
        os.remove(zip_path)
        
        print(f'Backup CSV subido por SFTP como {nombre_archivo}.')
        
    except Exception as e:
        print(f'Error al exportar CSV por SFTP: {e}')

@app.route('/exportar_csv')
@login_required
def exportar_csv():
    try:
        # Crear archivo ZIP con todos los CSV en directorio temporal
        import zipfile
        import tempfile
        fecha = datetime.now().strftime('%Y%m%d')
        nombre_archivo = f'lvm_csv_{fecha}.zip'
        
        # Crear directorio temporal
        with tempfile.TemporaryDirectory() as temp_dir:
            zip_path = os.path.join(temp_dir, nombre_archivo)
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Exportar cada tabla a CSV
                tablas = [Cuenta, Movimiento, MovimientoConcepto]
                nombres_tablas = ['cuentas', 'movimientos', 'movimientos_conceptos']
                
                for tabla, nombre in zip(tablas, nombres_tablas):
                    csv_buffer = io.StringIO()
                    writer = csv.writer(csv_buffer)
                    
                    # Obtener datos de la tabla
                    registros = tabla.query.all()
                    
                    if registros:
                        # Escribir encabezados
                        columnas = [column.name for column in tabla.__table__.columns]
                        writer.writerow(columnas)
                        
                        # Escribir datos
                        for registro in registros:
                            fila = []
                            for columna in columnas:
                                valor = getattr(registro, columna)
                                fila.append(str(valor) if valor is not None else '')
                            writer.writerow(fila)
                    
                    # Añadir CSV al ZIP
                    zipf.writestr(f'{nombre}.csv', csv_buffer.getvalue())
                
            # Subir al FTP usando el archivo temporal
            exportar_csv_a_ftp_from_path(zip_path, nombre_archivo)
            
            # Devolver archivo para descarga (se eliminará automáticamente al salir del contexto)
            return send_file(zip_path, as_attachment=True, download_name=nombre_archivo)
        
    except Exception as e:
        flash(f'Error al exportar CSV: {e}', 'error')
        return redirect(url_for('index'))

def exportar_csv_a_ftp_from_path(zip_path, nombre_archivo):
    sftp_host = os.environ.get('FTP_HOST')
    sftp_user = os.environ.get('FTP_USER')
    sftp_pass = os.environ.get('FTP_PASS')
    sftp_dir = os.environ.get('FTP_DIR', '/')
    
    if not sftp_host or not sftp_user or not sftp_pass:
        print('Faltan variables de entorno para la conexión SFTP.')
        return
    
    try:
        # Subir ZIP al FTP
        transport = paramiko.Transport((sftp_host, 22))
        transport.connect(username=sftp_user, password=sftp_pass)
        sftp = paramiko.SFTPClient.from_transport(transport)
        
        try:
            sftp.chdir(sftp_dir)
        except IOError:
            dirs = sftp_dir.strip('/').split('/')
            path = ''
            for d in dirs:
                path += '/' + d
                try:
                    sftp.chdir(path)
                except IOError:
                    sftp.mkdir(path)
                    sftp.chdir(path)
        
        # Listar archivos de backup CSV existentes
        archivos = sftp.listdir()
        backups_csv = sorted([f for f in archivos if f.startswith('lvm_csv_') and f.endswith('.zip')])
        # Si hay más de 2, borrar los más antiguos
        if len(backups_csv) > 2:
            for f in backups_csv[:-2]:
                try:
                    sftp.remove(f)
                    print(f'Backup CSV antiguo eliminado: {f}')
                except Exception as e:
                    print(f'No se pudo eliminar {f}: {e}')
        
        # Subir el nuevo backup CSV
        sftp.put(zip_path, nombre_archivo)
        sftp.close()
        transport.close()
        
        print(f'Backup CSV subido por SFTP como {nombre_archivo}.')
        
    except Exception as e:
        print(f'Error al exportar CSV por SFTP: {e}')

@app.route('/importar_csv', methods=['GET', 'POST'])
@login_required
def importar_csv():
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo.', 'error')
            return redirect(url_for('importar_csv'))
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            flash('No se seleccionó ningún archivo.', 'error')
            return redirect(url_for('importar_csv'))
        
        if not archivo.filename.endswith('.zip'):
            flash('El archivo debe ser un ZIP.', 'error')
            return redirect(url_for('importar_csv'))
        
        try:
            import zipfile
            import tempfile
            
            # Crear directorio temporal
            with tempfile.TemporaryDirectory() as temp_dir:
                # Guardar archivo ZIP
                zip_path = os.path.join(temp_dir, archivo.filename)
                archivo.save(zip_path)
                
                # Extraer y procesar CSV
                with zipfile.ZipFile(zip_path, 'r') as zipf:
                    # Mapeo de nombres de archivo a modelos
                    mapeo_tablas = {
                        'cuentas.csv': Cuenta,
                        'movimientos.csv': Movimiento,
                        'movimientos_conceptos.csv': MovimientoConcepto
                    }
                    
                    for nombre_archivo in zipf.namelist():
                        if nombre_archivo in mapeo_tablas:
                            modelo = mapeo_tablas[nombre_archivo]
                            
                            # Leer CSV
                            with zipf.open(nombre_archivo, 'r') as csv_file:
                                csv_reader = csv.reader(io.TextIOWrapper(csv_file, encoding='utf-8'))
                                encabezados = next(csv_reader)  # Saltar encabezados
                                
                                # Limpiar tabla existente
                                modelo.query.delete()
                                
                                # Insertar nuevos datos
                                for fila in csv_reader:
                                    if len(fila) == len(encabezados):
                                        registro = {}
                                        for i, columna in enumerate(encabezados):
                                            valor = fila[i]
                                            if valor == '':
                                                valor = None
                                            elif columna in ['id', 'cuenta_id', 'movimiento_id', 'contrapartida_id', 'cuenta_asociada_id']:
                                                try:
                                                    valor = int(valor) if valor else None
                                                except ValueError:
                                                    valor = None
                                            elif columna in ['base_imponible', 'total', 'importe']:
                                                try:
                                                    valor = float(valor) if valor else 0.0
                                                except ValueError:
                                                    valor = 0.0
                                            
                                            registro[columna] = valor
                                        
                                        nuevo_registro = modelo(**registro)
                                        db.session.add(nuevo_registro)
                
                db.session.commit()
                flash('Datos importados correctamente.', 'success')
                
        except Exception as e:
            flash(f'Error al importar CSV: {e}', 'error')
            db.session.rollback()
        
        return redirect(url_for('index'))
    
    return render_template('importar_csv.html')

@app.route('/extraer_nominas_pdf', methods=['GET', 'POST'])
def extraer_nominas_pdf():
    if request.method == 'GET':
        return render_template('extraer_nominas.html')
    
    if 'pdf_file' not in request.files:
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(request.url)
    
    file = request.files['pdf_file']
    if file.filename == '':
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(request.url)
    
    if file and file.filename.endswith('.pdf'):
        try:
            # Guardar el archivo subido
            pdf_path = os.path.join(app.root_path, file.filename)
            file.save(pdf_path)
            print(f"PDF guardado en: {pdf_path}")
            
            # Extraer texto del PDF
            texto_completo = ""
            with open(pdf_path, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    texto_completo += page.extract_text() + "\n"
            
            print(f"Texto extraído: {len(texto_completo)} caracteres")
            
            # Extraer datos específicos
            datos_nomina = extraer_datos_nomina(texto_completo)
            
            if not datos_nomina:
                flash('No se pudieron extraer datos del PDF', 'error')
                return redirect(request.url)
            
            # Crear Excel con los datos
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos Nóminas"
            
            # Encabezados
            headers = ['Nombre Trabajador', 'Total Devengado', 'Total Aportaciones', 'IRPF', 'Líquido a Percibir', 'Total SS Empresa', 'Archivo PDF']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Datos
            row = 2
            ws.cell(row=row, column=1, value=datos_nomina.get('nombre_trabajador', ''))
            ws.cell(row=row, column=2, value=datos_nomina.get('total_devengado', ''))
            ws.cell(row=row, column=3, value=datos_nomina.get('total_aportaciones', ''))
            ws.cell(row=row, column=4, value=datos_nomina.get('irpf', ''))
            ws.cell(row=row, column=5, value=datos_nomina.get('liquido_percibir', ''))
            ws.cell(row=row, column=6, value=datos_nomina.get('total_ss_empresa', ''))
            ws.cell(row=row, column=7, value=file.filename)
            
            # Guardar Excel
            excel_path_temp = os.path.join(app.root_path, f'nomina_datos_{datetime.now().strftime("%Y%m%d")}.xlsx')
            wb.save(excel_path_temp)
            wb.close()
            
            # Copiar a ubicación final
            excel_path_final = os.path.join(app.root_path, f'nomina_datos_{datetime.now().strftime("%Y%m%d")}.xlsx')
            shutil.copy2(excel_path_temp, excel_path_final)
            
            print(f"Excel guardado en: {excel_path_final}")
            
            # Limpiar archivo temporal
            os.remove(pdf_path)
            os.remove(excel_path_temp)
            
            flash('Datos extraídos correctamente', 'success')
            
            # Enviar archivo al usuario
            response = send_file(
                excel_path_final,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'nomina_datos_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )
            
            @response.call_on_close
            def cleanup():
                try:
                    os.remove(excel_path_final)
                except:
                    pass
            
            return response
            
        except Exception as e:
            flash(f'Error al extraer datos de nóminas: {str(e)}', 'error')
            return redirect(request.url)
    else:
        flash('Por favor selecciona un archivo PDF válido', 'error')
        return redirect(request.url)

def extraer_datos_nomina(texto):
    """Extrae datos específicos del texto de la nómina"""
    datos = {}
    
    # Dividir el texto en líneas para análisis más preciso
    lineas = texto.split('\n')
    
    # Buscar nombres de trabajadores
    for i, linea in enumerate(lineas):
        if 'LOGISTICA VENANCIO MATEOS SL' in linea:
            # Buscar en las siguientes líneas el nombre
            for j in range(i+1, min(i+5, len(lineas))):
                nombre_match = re.search(r'([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)*)', lineas[j])
                if nombre_match:
                    datos['nombre_trabajador'] = nombre_match.group(1).strip()
                    print(f"Encontrado nombre_trabajador: {datos['nombre_trabajador']}")
                    break
            break
    
    # Buscar total devengado - está en la línea que contiene "A. TOTAL DEVENGADO"
    for linea in lineas:
        if 'A. TOTAL DEVENGADO' in linea:
            # Buscar números en esa línea
            numeros = re.findall(r'([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)', linea)
            if numeros:
                datos['total_devengado'] = numeros[-1].replace(',', '.')  # Tomar el último número
                print(f"Encontrado total_devengado: {datos['total_devengado']}")
            break
    
    # Buscar total aportaciones - está en la línea que contiene "1-.TOTAL APORTACIONES"
    for linea in lineas:
        if '1-.TOTAL APORTACIONES' in linea:
            numeros = re.findall(r'([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)', linea)
            if numeros:
                datos['total_aportaciones'] = numeros[-1].replace(',', '.')
                print(f"Encontrado total_aportaciones: {datos['total_aportaciones']}")
            break
    
    # Buscar IRPF - está en la línea que contiene "2-. I.R.P.F"
    for linea in lineas:
        if '2-. I.R.P.F' in linea:
            numeros = re.findall(r'([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)', linea)
            if numeros:
                datos['irpf'] = numeros[-1].replace(',', '.')
                print(f"Encontrado irpf: {datos['irpf']}")
            break
    
    # Buscar líquido a percibir - está en la línea que contiene "LIQUIDO TOTAL A PERCIBIR"
    for linea in lineas:
        if 'LIQUIDO TOTAL A PERCIBIR' in linea:
            numeros = re.findall(r'([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)', linea)
            if numeros:
                datos['liquido_percibir'] = numeros[-1].replace(',', '.')
                print(f"Encontrado liquido_percibir: {datos['liquido_percibir']}")
            break
    
    # Buscar total SS empresa - está en la línea que contiene "Total SS Empresa"
    for linea in lineas:
        if 'Total SS Empresa' in linea:
            numeros = re.findall(r'([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)', linea)
            if numeros:
                datos['total_ss_empresa'] = numeros[-1].replace(',', '.')
                print(f"Encontrado total_ss_empresa: {datos['total_ss_empresa']}")
            break
    
    # Si no encuentra los datos con los patrones específicos, buscar de forma más general
    if 'total_devengado' not in datos:
        # Buscar números grandes que podrían ser el total devengado
        devengados_general = re.findall(r'([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)\s*\.{10,}', texto)
        if devengados_general:
            datos['total_devengado'] = devengados_general[0].replace(',', '.')
            print(f"Encontrado total_devengado (general): {datos['total_devengado']}")
    
    if 'total_aportaciones' not in datos:
        # Buscar números después de "TOTAL APORTACIONES"
        aportaciones_general = re.findall(r'TOTAL\s*APORTACIONES[^\d]*([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)', texto)
        if aportaciones_general:
            datos['total_aportaciones'] = aportaciones_general[0].replace(',', '.')
            print(f"Encontrado total_aportaciones (general): {datos['total_aportaciones']}")
    
    if 'irpf' not in datos:
        # Buscar números después de "I.R.P.F"
        irpf_general = re.findall(r'I\.R\.P\.F[^\d]*([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)', texto)
        if irpf_general:
            datos['irpf'] = irpf_general[0].replace(',', '.')
            print(f"Encontrado irpf (general): {datos['irpf']}")
    
    if 'liquido_percibir' not in datos:
        # Buscar números después de "A PERCIBIR"
        liquidos_general = re.findall(r'A\s*PERCIBIR[^\d]*([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)', texto)
        if liquidos_general:
            datos['liquido_percibir'] = liquidos_general[0].replace(',', '.')
            print(f"Encontrado liquido_percibir (general): {datos['liquido_percibir']}")
    
    if 'total_ss_empresa' not in datos:
        # Buscar números después de "SS Empresa"
        ss_general = re.findall(r'SS\s*Empresa[^\d]*([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2})?)', texto)
        if ss_general:
            datos['total_ss_empresa'] = ss_general[0].replace(',', '.')
            print(f"Encontrado total_ss_empresa (general): {datos['total_ss_empresa']}")
    
    return datos

@app.route('/test_pdf_extraction')
def test_pdf_extraction():
    """Función de prueba para extraer texto del PDF"""
    pdf_path = os.path.join(app.root_path, 'LOGISTICA VENANCIO JULIO.pdf')
    
    if not os.path.exists(pdf_path):
        return f"<h3>Error: No se encontró el archivo PDF en {pdf_path}</h3>"
    
    try:
        texto_completo = ""
        with open(pdf_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                texto_pagina = page.extract_text()
                texto_completo += f"--- PÁGINA {page_num + 1} ---\n{texto_pagina}\n\n"
        
        return f"""
        <h3>Texto extraído del PDF</h3>
        <p><strong>Total de caracteres extraídos:</strong> {len(texto_completo)}</p>
        <p><strong>Archivo:</strong> LOGISTICA VENANCIO JULIO.pdf</p>
        <p><strong>Número de páginas:</strong> {len(pdf_reader.pages)}</p>
        <hr>
        <pre style="max-height: 500px; overflow-y: scroll; border: 1px solid #ccc; padding: 10px; background-color: #f9f9f9;">{texto_completo}</pre>
        """
        
    except Exception as e:
        return f"<h3>Error al extraer texto del PDF:</h3><p>{str(e)}</p>"

@app.route('/guardar_cambios', methods=['POST'])
@login_required
def guardar_cambios():
    subir_db_a_ftp()
    flash('Cambios guardados y base de datos subida correctamente.', 'success')
    return redirect(request.referrer or url_for('index'))

@app.template_filter('tipomov')
def tipomov(value):
    if value is None:
        return ''
    if value.lower() == 'gasto':
        return 'Gasto'
    elif value.lower() == 'ingreso':
        return 'Ingreso'
    return value.capitalize()

@app.template_filter('datetimeformat')
def datetimeformat(value, format='%d/%m/%Y'):
    try:
        from datetime import datetime
        return datetime.strptime(value, '%Y-%m-%d').strftime(format)
    except Exception:
        return value

@app.template_filter('dateinput')
def dateinput(value):
    """Convierte fecha de dd/mm/yyyy a yyyy-mm-dd para input type='date'"""
    try:
        from datetime import datetime
        if value:
            # Si ya está en formato yyyy-mm-dd, devolverlo tal como está
            if '-' in value and len(value.split('-')[0]) == 4:
                return value
            # Si está en formato dd/mm/yyyy, convertirlo
            return datetime.strptime(value, '%d/%m/%Y').strftime('%Y-%m-%d')
    except Exception:
        pass
    return value

def convertir_fechas_para_filtro(fecha_inicio, fecha_fin):
    """
    Convierte fechas de YYYY-MM-DD a objetos datetime para usar en filtros de base de datos
    """
    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
    fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
    return fecha_inicio_dt, fecha_fin_dt

def parsear_fecha_robusto(fecha_str):
    """
    Intenta parsear una fecha en diferentes formatos comunes
    """
    if not fecha_str:
        return None
    
    # Lista de formatos a intentar
    formatos = [
        '%d/%m/%Y',    # 01/01/2024
        '%d/%m/%y',    # 01/01/24
        '%d-%m-%Y',    # 01-01-2024
        '%d-%m-%y',    # 01-01-24
        '%Y-%m-%d',    # 2024-01-01
        '%Y/%m/%d',    # 2024/01/01
        '%d/%m/%Y',    # 1/1/2024 (con ceros a la izquierda)
        '%d/%m/%y',    # 1/1/24 (con ceros a la izquierda)
    ]
    
    for formato in formatos:
        try:
            return datetime.strptime(fecha_str, formato)
        except ValueError:
            continue
    
    # Si ninguno funciona, intentar limpiar la fecha
    fecha_limpia = fecha_str.strip()
    if fecha_limpia != fecha_str:
        return parsear_fecha_robusto(fecha_limpia)
    
    return None

@app.route('/configurar_general_nominas', methods=['GET', 'POST'])
@login_required
def configurar_general_nominas():
    if request.method == 'POST':
        # Convertir fechas de dd/mm/yyyy a yyyy-mm-dd antes de guardar
        fecha_trabajo = request.form['fecha_trabajo']
        fecha_factura = request.form['fecha_factura']
        
        # Convertir formato si viene en dd/mm/yyyy
        if '/' in fecha_trabajo and len(fecha_trabajo.split('/')) == 3:
            partes = fecha_trabajo.split('/')
            if len(partes[0]) == 2 and len(partes[1]) == 2 and len(partes[2]) == 4:
                fecha_trabajo = f"{partes[2]}-{partes[1]}-{partes[0]}"
        
        if '/' in fecha_factura and len(fecha_factura.split('/')) == 3:
            partes = fecha_factura.split('/')
            if len(partes[0]) == 2 and len(partes[1]) == 2 and len(partes[2]) == 4:
                fecha_factura = f"{partes[2]}-{partes[1]}-{partes[0]}"
        
        # Guardar configuración general
        session['nomina_general_config'] = {
            'fecha_trabajo': fecha_trabajo,
            'fecha_factura': fecha_factura,
            'num_factura': request.form['num_factura']
        }
        flash('Configuración general de nóminas guardada correctamente.', 'success')
        return redirect(url_for('ver_todas_nominas'))
    
    # Cargar configuración existente o valores por defecto
    config = session.get('nomina_general_config', {
        'fecha_trabajo': datetime.now().strftime('%Y-%m-%d'),
        'fecha_factura': datetime.now().strftime('%Y-%m-%d'),
        'num_factura': f'Nómina {datetime.now().strftime("%B %Y")}'
    })
    
    # Convertir fechas a formato dd/mm/yyyy para mostrar en la interfaz
    if '-' in config['fecha_trabajo'] and len(config['fecha_trabajo'].split('-')) == 3:
        partes = config['fecha_trabajo'].split('-')
        if len(partes[0]) == 4 and len(partes[1]) == 2 and len(partes[2]) == 2:
            config['fecha_trabajo'] = f"{partes[2]}/{partes[1]}/{partes[0]}"
    
    if '-' in config['fecha_factura'] and len(config['fecha_factura'].split('-')) == 3:
        partes = config['fecha_factura'].split('-')
        if len(partes[0]) == 4 and len(partes[1]) == 2 and len(partes[2]) == 2:
            config['fecha_factura'] = f"{partes[2]}/{partes[1]}/{partes[0]}"
    
    return render_template('configurar_general_nominas.html', config=config)

@app.route('/generar_todas_nominas', methods=['POST'])
@login_required
def generar_todas_nominas():
    # Obtener configuración general
    general_config = session.get('nomina_general_config')
    if not general_config:
        flash('Debe configurar los parámetros generales primero.', 'error')
        return redirect(url_for('configurar_general_nominas'))
    
    # Obtener todos los empleados
    empleados = Cuenta.query.filter(
        Cuenta.tipo == 'contrapartida',
        Cuenta.nombre.like('EMP%')
    ).all()
    
    if not empleados:
        flash('No se encontraron empleados en la base de datos.', 'error')
        return redirect(url_for('configurar_general_nominas'))
    
    # Obtener cuentas necesarias
    cuentas = {
        'sueldos': Cuenta.query.filter_by(cuenta='640000000001').first(),
        'retencion': Cuenta.query.filter_by(cuenta='47510000001').first(),
        'ss_trabajador': Cuenta.query.filter_by(cuenta='642000000002').first(),
        'ss_empresa': Cuenta.query.filter_by(cuenta='642000000001').first(),
        'dietas': Cuenta.query.filter_by(cuenta='649000000002').first()
    }
    
    # Crear cuentas faltantes automáticamente
    cuentas_por_crear = {
        'sueldos': ('640000000001', 'SUELDOS Y SALARIOS'),
        'dietas': ('649000000002', 'DIETAS TRABAJADORES')
    }
    
    for nombre, (numero, nombre_cuenta) in cuentas_por_crear.items():
        if not cuentas[nombre]:
            nueva_cuenta = Cuenta(
                cuenta=numero,
                nombre=nombre_cuenta,
                tipo='normal'
            )
            db.session.add(nueva_cuenta)
            db.session.flush()
            cuentas[nombre] = nueva_cuenta
            print(f"Cuenta creada: {numero} - {nombre_cuenta}")
    
    # Commit para guardar las cuentas creadas
    db.session.commit()
    
    # Verificar que todas las cuentas existen
    cuentas_faltantes = [k for k, v in cuentas.items() if not v]
    if cuentas_faltantes:
        flash(f'Faltan las siguientes cuentas: {", ".join(cuentas_faltantes)}', 'error')
        return redirect(url_for('configurar_general_nominas'))
    

    
    movimientos_creados = 0
    
    for empleado in empleados:
        # Buscar la última nómina del empleado buscando movimientos donde él sea la contrapartida
        ultima_nomina = db.session.query(Movimiento).join(MovimientoConcepto).filter(
            MovimientoConcepto.contrapartida_id == empleado.id,
            Movimiento.tipo == 'Gasto',
            Movimiento.num_factura.like('Nómina%')
        ).order_by(Movimiento.fecha_trabajo.desc()).first()
        
        # Valores por defecto: 0 si no hay nómina previa, valores de la última nómina si existe
        valores_default = {
            'liquido_percibir': 0.00,
            'retencion_irpf': 0.00,
            'ss_trabajador': 0.00,
            'ss_empresa': 0.00,
            'dietas': 0.00
        }
        
        # Si existe una nómina anterior, usar esos valores
        if ultima_nomina:
            conceptos = MovimientoConcepto.query.filter_by(movimiento_id=ultima_nomina.id).all()
            
            for concepto in conceptos:
                if concepto.cuenta.cuenta == '640000000001':  # Sueldos
                    valores_default['liquido_percibir'] = concepto.importe
                elif concepto.cuenta.cuenta == '47510000001':  # Retención
                    valores_default['retencion_irpf'] = concepto.importe
                elif concepto.cuenta.cuenta == '642000000002':  # SS Trabajador
                    valores_default['ss_trabajador'] = concepto.importe
                elif concepto.cuenta.cuenta == '642000000001':  # SS Empresa
                    valores_default['ss_empresa'] = concepto.importe
                elif concepto.cuenta.cuenta == '649000000002':  # Dietas
                    valores_default['dietas'] = concepto.importe
        
        # Obtener configuración específica del empleado
        config_key = f'nomina_config_{empleado.id}'
        empleado_config = session.get(config_key, valores_default)
        
        # Calcular base imponible y total
        base_imponible = empleado_config['liquido_percibir'] + empleado_config['ss_trabajador'] + empleado_config['retencion_irpf']
        total = empleado_config['liquido_percibir'] + empleado_config['ss_empresa'] + empleado_config['dietas']
        
        # Crear movimiento principal
        # Añadir el nombre del empleado al número de factura para evitar confusión
        num_factura_empleado = f"{general_config['num_factura']} - {empleado.nombre}"
        movimiento = Movimiento(
            tipo='Gasto',
            fecha_trabajo=general_config['fecha_trabajo'],
            fecha_factura=general_config['fecha_factura'],
            num_factura=num_factura_empleado,
            base_imponible=base_imponible,
            total=total
        )
        db.session.add(movimiento)
        db.session.flush()  # Para obtener el ID del movimiento
        
        # Crear conceptos del movimiento
        conceptos = [
            (cuentas['sueldos'], empleado_config['liquido_percibir']),
            (cuentas['retencion'], empleado_config['retencion_irpf']),
            (cuentas['ss_trabajador'], empleado_config['ss_trabajador']),
            (cuentas['ss_empresa'], empleado_config['ss_empresa']),
            (cuentas['dietas'], empleado_config['dietas'])
        ]
        
        for cuenta, importe in conceptos:
            concepto = MovimientoConcepto(
                movimiento_id=movimiento.id,
                cuenta_id=cuenta.id,
                contrapartida_id=empleado.id,
                importe=importe,
                concepto=''
            )
            db.session.add(concepto)
        
        movimientos_creados += 1
        
        # Limpiar configuración específica del empleado
        session.pop(config_key, None)
    
    try:
        db.session.commit()
        flash(f'Se han creado {movimientos_creados} movimientos de nómina correctamente.', 'success')
        # Limpiar configuración general
        session.pop('nomina_general_config', None)
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear los movimientos: {str(e)}', 'error')
    
    return redirect(url_for('listar_movimientos'))

@app.route('/ver_todas_nominas')
@login_required
def ver_todas_nominas():
    # Obtener configuración general
    general_config = session.get('nomina_general_config', {
        'fecha_trabajo': datetime.now().strftime('%Y-%m-%d'),
        'fecha_factura': datetime.now().strftime('%Y-%m-%d'),
        'num_factura': f'Nómina {datetime.now().strftime("%B %Y")}'
    })
    
    # Convertir fechas a formato dd/mm/yyyy para mostrar en la interfaz
    config_display = general_config.copy()
    if '-' in config_display['fecha_trabajo'] and len(config_display['fecha_trabajo'].split('-')) == 3:
        partes = config_display['fecha_trabajo'].split('-')
        if len(partes[0]) == 4 and len(partes[1]) == 2 and len(partes[2]) == 2:
            config_display['fecha_trabajo'] = f"{partes[2]}/{partes[1]}/{partes[0]}"
    
    if '-' in config_display['fecha_factura'] and len(config_display['fecha_factura'].split('-')) == 3:
        partes = config_display['fecha_factura'].split('-')
        if len(partes[0]) == 4 and len(partes[1]) == 2 and len(partes[2]) == 2:
            config_display['fecha_factura'] = f"{partes[2]}/{partes[1]}/{partes[0]}"
    
    # Obtener todos los empleados
    empleados = Cuenta.query.filter(
        Cuenta.tipo == 'contrapartida',
        Cuenta.nombre.like('EMP%')
    ).order_by(Cuenta.nombre).all()
    
    # Obtener configuraciones de cada empleado
    empleados_config = []
    for empleado in empleados:
        # Buscar la última nómina del empleado buscando movimientos donde él sea la contrapartida
        ultima_nomina = db.session.query(Movimiento).join(MovimientoConcepto).filter(
            MovimientoConcepto.contrapartida_id == empleado.id,
            Movimiento.tipo == 'Gasto',
            Movimiento.num_factura.like('Nómina%')
        ).order_by(Movimiento.fecha_trabajo.desc()).first()
        
        # Valores por defecto: 0 si no hay nómina previa, valores de la última nómina si existe
        valores_default = {
            'liquido_percibir': 0.00,
            'retencion_irpf': 0.00,
            'ss_trabajador': 0.00,
            'ss_empresa': 0.00,
            'dietas': 0.00
        }
        
        # Si existe una nómina anterior, usar esos valores
        if ultima_nomina:
            conceptos = MovimientoConcepto.query.filter_by(movimiento_id=ultima_nomina.id).all()
            print(f"DEBUG - Empleado: {empleado.nombre} - Última nómina: {ultima_nomina.num_factura}")
            
            for concepto in conceptos:
                print(f"DEBUG - Concepto: Cuenta {concepto.cuenta.cuenta} - {concepto.cuenta.nombre}, Importe: {concepto.importe}")
                if concepto.cuenta.cuenta == '640000000001':  # Sueldos
                    valores_default['liquido_percibir'] = concepto.importe
                    print(f"DEBUG - Asignado liquido_percibir: {concepto.importe}")
                elif concepto.cuenta.cuenta == '47510000001':  # Retención
                    valores_default['retencion_irpf'] = concepto.importe
                    print(f"DEBUG - Asignado retencion_irpf: {concepto.importe}")
                elif concepto.cuenta.cuenta == '642000000002':  # SS Trabajador
                    valores_default['ss_trabajador'] = concepto.importe
                    print(f"DEBUG - Asignado ss_trabajador: {concepto.importe}")
                elif concepto.cuenta.cuenta == '642000000001':  # SS Empresa
                    valores_default['ss_empresa'] = concepto.importe
                    print(f"DEBUG - Asignado ss_empresa: {concepto.importe}")
                elif concepto.cuenta.cuenta == '649000000002':  # Dietas
                    valores_default['dietas'] = concepto.importe
                    print(f"DEBUG - Asignado dietas: {concepto.importe}")
                else:
                    print(f"DEBUG - Cuenta NO reconocida: {concepto.cuenta.cuenta}")
            
            print(f"DEBUG - Valores finales para {empleado.nombre}: {valores_default}")
        else:
            print(f"DEBUG - Empleado: {empleado.nombre} - No se encontró ninguna nómina anterior")
        
        # Obtener configuración actual o usar valores de la última nómina
        config_key = f'nomina_config_{empleado.id}'
        empleado_config = session.get(config_key, valores_default)
        
        # Calcular total
        total = empleado_config['liquido_percibir'] + empleado_config['ss_empresa'] + empleado_config['dietas'] + empleado_config['retencion_irpf'] + empleado_config['ss_trabajador']
        
        empleados_config.append({
            'empleado': empleado,
            'config': empleado_config,
            'total': total,
            'tiene_ultima_nomina': ultima_nomina is not None
        })
    
    return render_template('ver_todas_nominas.html', empleados_config=empleados_config, general_config=config_display)

@app.route('/guardar_config_empleado/<int:empleado_id>', methods=['POST'])
@login_required
def guardar_config_empleado(empleado_id):
    empleado = Cuenta.query.get_or_404(empleado_id)
    
    # Guardar configuración específica del empleado
    config_key = f'nomina_config_{empleado_id}'
    session[config_key] = {
        'liquido_percibir': float(request.form['liquido_percibir']),
        'retencion_irpf': float(request.form['retencion_irpf']),
        'ss_trabajador': float(request.form['ss_trabajador']),
        'ss_empresa': float(request.form['ss_empresa']),
        'dietas': float(request.form['dietas'])
    }
    
    flash(f'Configuración de {empleado.nombre} guardada correctamente.', 'success')
    return redirect(url_for('ver_todas_nominas'))

# ==================== CONTROL DE GASOIL ====================

@app.route('/control_gasoil')
@login_required
def control_gasoil():
    """Página principal del control de gasoil"""
    vehiculos = Vehiculo.query.filter_by(activo=True).all()
    return render_template('control_gasoil.html', vehiculos=vehiculos)

# Gestión de vehículos
@app.route('/control_gasoil/vehiculos')
@login_required
def listar_vehiculos():
    """Listar todos los vehículos"""
    vehiculos = Vehiculo.query.order_by(Vehiculo.matricula).all()
    return render_template('vehiculos.html', vehiculos=vehiculos)

@app.route('/control_gasoil/vehiculos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_vehiculo():
    """Crear nuevo vehículo"""
    if request.method == 'POST':
        matricula = request.form['matricula'].upper().strip()
        marca = request.form['marca'].strip()
        modelo = request.form['modelo'].strip()
        año_compra = int(request.form['año_compra'])
        observaciones = request.form.get('observaciones', '').strip()
        
        # Verificar que la matrícula no esté duplicada
        existe = Vehiculo.query.filter_by(matricula=matricula).first()
        if existe:
            flash('Ya existe un vehículo con esa matrícula.', 'error')
            return render_template('vehiculo_form.html')
        
        nuevo_vehiculo = Vehiculo(
            matricula=matricula,
            marca=marca,
            modelo=modelo,
            año_compra=año_compra,
            observaciones=observaciones
        )
        
        db.session.add(nuevo_vehiculo)
        db.session.commit()
        flash('Vehículo creado correctamente.', 'success')
        return redirect(url_for('listar_vehiculos'))
    
    return render_template('vehiculo_form.html')

@app.route('/control_gasoil/vehiculos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_vehiculo(id):
    """Editar vehículo existente"""
    vehiculo = Vehiculo.query.get_or_404(id)
    
    if request.method == 'POST':
        matricula = request.form['matricula'].upper().strip()
        marca = request.form['marca'].strip()
        modelo = request.form['modelo'].strip()
        año_compra = int(request.form['año_compra'])
        observaciones = request.form.get('observaciones', '').strip()
        activo = 'activo' in request.form
        
        # Verificar que la matrícula no esté duplicada (excepto el propio vehículo)
        existe = Vehiculo.query.filter(Vehiculo.matricula == matricula, Vehiculo.id != id).first()
        if existe:
            flash('Ya existe un vehículo con esa matrícula.', 'error')
            return render_template('vehiculo_form.html', vehiculo=vehiculo)
        
        vehiculo.matricula = matricula
        vehiculo.marca = marca
        vehiculo.modelo = modelo
        vehiculo.año_compra = año_compra
        vehiculo.observaciones = observaciones
        vehiculo.activo = activo
        
        db.session.commit()
        flash('Vehículo actualizado correctamente.', 'success')
        return redirect(url_for('listar_vehiculos'))
    
    return render_template('vehiculo_form.html', vehiculo=vehiculo)

@app.route('/control_gasoil/vehiculos/borrar/<int:id>', methods=['POST'])
@login_required
def borrar_vehiculo(id):
    """Borrar vehículo (marcar como inactivo)"""
    vehiculo = Vehiculo.query.get_or_404(id)
    
    # Verificar si tiene consumos asociados
    if vehiculo.consumos:
        flash('No se puede borrar el vehículo porque tiene consumos asociados. Se marcará como inactivo.', 'warning')
        vehiculo.activo = False
        db.session.commit()
    else:
        db.session.delete(vehiculo)
        db.session.commit()
        flash('Vehículo borrado correctamente.', 'success')
    
    return redirect(url_for('listar_vehiculos'))

# Gestión de consumos
@app.route('/control_gasoil/consumos')
@login_required
def listar_consumos():
    """Listar todos los consumos"""
    vehiculo_id = request.args.get('vehiculo_id', type=int)
    año = request.args.get('año', type=int)
    mes = request.args.get('mes', type=int)
    
    query = ConsumoGasoil.query.join(Vehiculo)
    
    if vehiculo_id:
        query = query.filter(ConsumoGasoil.vehiculo_id == vehiculo_id)
    if año:
        query = query.filter(db.func.strftime('%Y', ConsumoGasoil.fecha) == str(año))
    if mes:
        query = query.filter(db.func.strftime('%m', ConsumoGasoil.fecha) == f"{mes:02d}")
    
    consumos = query.order_by(ConsumoGasoil.fecha.desc()).all()
    vehiculos = Vehiculo.query.filter_by(activo=True).all()
    
    return render_template('consumos.html', consumos=consumos, vehiculos=vehiculos, 
                         vehiculo_seleccionado=vehiculo_id, año_seleccionado=año, mes_seleccionado=mes)

@app.route('/control_gasoil/consumos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_consumo():
    """Crear nuevo consumo"""
    vehiculos = Vehiculo.query.filter_by(activo=True).all()
    
    if request.method == 'POST':
        vehiculo_id = int(request.form['vehiculo_id'])
        fecha = request.form['fecha']
        litros = float(request.form['litros'])
        precio_total = float(request.form['precio_total'])
        kms = float(request.form['kms'])
        facturacion = float(request.form.get('facturacion', 0)) if request.form.get('facturacion') else None
        observaciones = request.form.get('observaciones', '').strip()
        
        # Convertir fecha de dd/mm/yyyy a yyyy-mm-dd si es necesario
        if '/' in fecha and len(fecha.split('/')) == 3:
            partes = fecha.split('/')
            if len(partes[0]) == 2 and len(partes[1]) == 2 and len(partes[2]) == 4:
                fecha = f"{partes[2]}-{partes[1]}-{partes[0]}"
        
        # Calcular precio por litro
        precio_litro = precio_total / litros if litros > 0 else 0
        total = precio_total
        
        nuevo_consumo = ConsumoGasoil(
            vehiculo_id=vehiculo_id,
            fecha=fecha,
            litros=litros,
            precio_litro=precio_litro,
            total=total,
            kms=kms,
            facturacion=facturacion,
            observaciones=observaciones
        )
        
        db.session.add(nuevo_consumo)
        db.session.commit()
        flash('Consumo registrado correctamente.', 'success')
        return redirect(url_for('listar_consumos'))
    
    return render_template('consumo_form.html', vehiculos=vehiculos)

@app.route('/control_gasoil/consumos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_consumo(id):
    """Editar consumo existente"""
    consumo = ConsumoGasoil.query.get_or_404(id)
    vehiculos = Vehiculo.query.filter_by(activo=True).all()
    
    if request.method == 'POST':
        vehiculo_id = int(request.form['vehiculo_id'])
        fecha = request.form['fecha']
        litros = float(request.form['litros'])
        precio_total = float(request.form['precio_total'])
        kms = float(request.form['kms'])
        facturacion = float(request.form.get('facturacion', 0)) if request.form.get('facturacion') else None
        observaciones = request.form.get('observaciones', '').strip()
        
        # Convertir fecha de dd/mm/yyyy a yyyy-mm-dd si es necesario
        if '/' in fecha and len(fecha.split('/')) == 3:
            partes = fecha.split('/')
            if len(partes[0]) == 2 and len(partes[1]) == 2 and len(partes[2]) == 4:
                fecha = f"{partes[2]}-{partes[1]}-{partes[0]}"
        
        # Calcular precio por litro
        precio_litro = precio_total / litros if litros > 0 else 0
        total = precio_total
        
        consumo.vehiculo_id = vehiculo_id
        consumo.fecha = fecha
        consumo.litros = litros
        consumo.precio_litro = precio_litro
        consumo.total = total
        consumo.kms = kms
        consumo.facturacion = facturacion
        consumo.observaciones = observaciones
        
        db.session.commit()
        flash('Consumo actualizado correctamente.', 'success')
        return redirect(url_for('listar_consumos'))
    
    return render_template('consumo_form.html', consumo=consumo, vehiculos=vehiculos)

@app.route('/control_gasoil/consumos/borrar/<int:id>', methods=['POST'])
@login_required
def borrar_consumo(id):
    """Borrar consumo"""
    consumo = ConsumoGasoil.query.get_or_404(id)
    db.session.delete(consumo)
    db.session.commit()
    flash('Consumo borrado correctamente.', 'success')
    return redirect(url_for('listar_consumos'))

# Análisis y estadísticas
@app.route('/control_gasoil/analisis')
@login_required
def analisis_gasoil():
    """Página de análisis y estadísticas"""
    vehiculo_id = request.args.get('vehiculo_id', type=int)
    año = request.args.get('año', type=int) or datetime.now().year
    
    # Obtener datos para análisis
    query = ConsumoGasoil.query.join(Vehiculo)
    if vehiculo_id:
        query = query.filter(ConsumoGasoil.vehiculo_id == vehiculo_id)
    
    # Filtrar por año
    query = query.filter(db.func.strftime('%Y', ConsumoGasoil.fecha) == str(año))
    
    consumos = query.order_by(ConsumoGasoil.fecha).all()
    vehiculos = Vehiculo.query.filter_by(activo=True).all()
    
    # Calcular estadísticas
    if consumos:
        total_litros = sum(c.litros for c in consumos)
        total_gasto = sum(c.total for c in consumos)
        total_kms = sum(c.kms for c in consumos)
        total_facturacion = sum(float(c.facturacion) if c.facturacion and str(c.facturacion).replace('.', '').replace('-', '').isdigit() else 0 for c in consumos)
        precio_promedio = total_gasto / total_litros if total_litros > 0 else 0
        consumo_por_km = total_litros / total_kms if total_kms > 0 else 0
        rentabilidad = total_facturacion - total_gasto if total_facturacion else 0
        
        # Datos por mes
        datos_mensuales = {}
        for consumo in consumos:
            mes = datetime.strptime(consumo.fecha, '%Y-%m-%d').month
            if mes not in datos_mensuales:
                datos_mensuales[mes] = {'litros': 0, 'gasto': 0, 'kms': 0, 'facturacion': 0}
            datos_mensuales[mes]['litros'] += consumo.litros
            datos_mensuales[mes]['gasto'] += consumo.total
            datos_mensuales[mes]['kms'] += consumo.kms
            if consumo.facturacion:
                facturacion_val = float(consumo.facturacion) if str(consumo.facturacion).replace('.', '').replace('-', '').isdigit() else 0
                datos_mensuales[mes]['facturacion'] += facturacion_val
    else:
        total_litros = total_gasto = total_kms = total_facturacion = precio_promedio = consumo_por_km = rentabilidad = 0
        datos_mensuales = {}
    
    return render_template('analisis_gasoil.html', 
                         consumos=consumos,
                         vehiculos=vehiculos,
                         vehiculo_seleccionado=vehiculo_id,
                         año_seleccionado=año,
                         total_litros=total_litros,
                         total_gasto=total_gasto,
                         total_kms=total_kms,
                         total_facturacion=total_facturacion,
                         precio_promedio=precio_promedio,
                         consumo_por_km=consumo_por_km,
                         rentabilidad=rentabilidad,
                         datos_mensuales=datos_mensuales)

# Función para migrar datos existentes de facturación de string a float
def migrar_facturacion():
    """Migra los datos de facturación de string a float"""
    with app.app_context():
        consumos = ConsumoGasoil.query.all()
        for consumo in consumos:
            if consumo.facturacion and isinstance(consumo.facturacion, str):
                try:
                    # Intentar convertir a float
                    consumo.facturacion = float(consumo.facturacion)
                except (ValueError, TypeError):
                    # Si no se puede convertir, poner a None
                    consumo.facturacion = None
        db.session.commit()
        print("Migración de facturación completada")

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # Ejecutar migración de facturación
        migrar_facturacion()
    app.run(debug=True) 